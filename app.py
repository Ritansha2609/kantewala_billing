from flask import Flask, render_template, request
import sqlite3
from datetime import datetime
from num2words import num2words
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.jinja_env.globals.update(enumerate=enumerate)

# -------------------- DB INIT --------------------
def init_db():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number INTEGER,
            date TEXT,
            customer_name TEXT,
            customer_address TEXT,
            customer_state TEXT,
            customer_phone TEXT,
            product_names TEXT,
            product_qtys TEXT,
            product_amounts TEXT,
            apply_gst TEXT,
            total_amount REAL,
            cgst REAL,
            sgst REAL,
            net_amount REAL
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# -------------------- ROUTES --------------------

@app.route('/')
def form():
    return render_template("form.html")

@app.route('/confirm_gst', methods=['POST'])
def confirm_gst():
    all_data = request.form.to_dict(flat=False)
    return render_template('confirm_gst.html', all_data=all_data)

@app.route('/generate_bill', methods=['POST'])
def generate_bill():
    name = request.form.get('customer_name', 'N/A')
    address = request.form.get('customer_address', 'N/A')
    state = request.form.get('customer_state', 'N/A')
    phone = request.form.get('customer_phone', 'N/A')
    apply_gst = request.form.get('apply_gst', 'no')

    product_names = request.form.getlist('product_name[]')
    product_qtys = list(map(int, request.form.getlist('product_qty[]')))
    product_amounts = list(map(float, request.form.getlist('product_amount[]')))

    subtotal = sum(product_amounts)
    cgst = sgst = 0.0

    if apply_gst == 'yes':
        cgst = round(subtotal * 0.09, 2)
        sgst = round(subtotal * 0.09, 2)

    net_total = round(subtotal + cgst + sgst, 2)
    amount_words = num2words(net_total, to='cardinal', lang='en').title().replace(',', '') + " Only /-"

    # Generate invoice number
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("SELECT MAX(invoice_number) FROM bills")
    result = c.fetchone()
    last_invoice = result[0] if result[0] is not None else 999
    invoice_number = last_invoice + 1

    # Save to DB
    c.execute('''
        INSERT INTO bills (
            invoice_number, date, customer_name, customer_address, customer_state, customer_phone,
            product_names, product_qtys, product_amounts, apply_gst,
            total_amount, cgst, sgst, net_amount
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        invoice_number,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        name, address, state, phone,
        ','.join(product_names),
        ','.join(map(str, product_qtys)),
        ','.join(map(str, product_amounts)),
        apply_gst,
        subtotal, cgst, sgst, net_total
    ))
    conn.commit()
    conn.close()

    # -------------------- Save to Excel --------------------
    excel_file = "invoices.xlsx"

    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append([
            "Invoice No", "Date", "Customer Name", "Address", "State", "Phone",
            "Product Names", "Quantities", "Amounts", "CGST", "SGST", "Total"
        ])
    else:
        wb = load_workbook(excel_file)
        ws = wb.active

    ws.append([
        invoice_number, datetime.now().strftime("%d-%m-%Y %H:%M"),
        name, address, state, phone,
        ', '.join(product_names),
        ', '.join(map(str, product_qtys)),
        ', '.join(map(str, product_amounts)),
        cgst, sgst, net_total
    ])

    wb.save(excel_file)
    # ------------------------------------------------------

    return render_template("bill.html",
        name=name,
        address=address,
        state=state,
        phone=phone,
        product_data=zip(product_names, product_qtys, product_amounts),
        subtotal=subtotal,
        cgst=cgst,
        sgst=sgst,
        total=net_total,
        amount_words=amount_words,
        date=datetime.now().strftime("%d-%m-%Y %H:%M"),
        invoice_number=invoice_number,
        apply_gst=apply_gst
    )

if __name__ == '__main__':
    app.run(debug=True)
